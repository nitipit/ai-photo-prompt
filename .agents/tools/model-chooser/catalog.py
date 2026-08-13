#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.12"
# dependencies = [
#   "openpyxl>=3.1.5",
#   "cyclopts>=3.0.0",
# ]
# ///
"""Validate approved model profiles and generate the agent-readable catalog."""

from __future__ import annotations

import math
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Annotated, Any

from cyclopts import App, Parameter
from openpyxl import load_workbook

ARTIFICIAL_ANALYSIS_URL = "https://artificialanalysis.ai/"
DEFAULT_DATA_DIR = Path(".agents/model-choosing")
DEFAULT_WORKBOOK = "model-profiles.xlsx"
DEFAULT_CATALOG = "model-catalog.md"
DEFAULT_PROJECT_ROOT = Path.cwd()
REQUIRED_HEADERS = (
    "Provider",
    "Model",
    "Thinking",
    "Cost/task (USD)",
    "Time/task (min)",
    "Intelligence index",
)
THINKING_LEVELS = {"off", "minimal", "low", "medium", "high", "xhigh", "max"}
IDENTIFIER_PATTERN = re.compile(r"^[a-z0-9][a-z0-9._/-]*$")


class CatalogError(RuntimeError):
    """Raised when the approved model profile workbook is invalid."""


@dataclass(frozen=True)
class ModelProfile:
    """One approved runtime model and its comparative decision metrics."""

    provider: str
    model: str
    thinking: str
    cost_per_task: float
    time_per_task: float
    intelligence_index: float


def clean_text(value: Any, *, field: str, row: int) -> str:
    if value is None:
        raise CatalogError(f"row {row}: {field} is required")
    text = str(value).strip()
    if not text:
        raise CatalogError(f"row {row}: {field} is required")
    return text


def positive_number(value: Any, *, field: str, row: int) -> float:
    if isinstance(value, bool):
        raise CatalogError(f"row {row}: {field} must be a positive number")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise CatalogError(f"row {row}: {field} must be a positive number") from exc
    if not math.isfinite(number) or number <= 0:
        raise CatalogError(f"row {row}: {field} must be a positive number")
    return number


def validate_identifier(value: str, *, field: str, row: int) -> str:
    if not IDENTIFIER_PATTERN.fullmatch(value):
        raise CatalogError(
            f"row {row}: {field} must be a lowercase runtime identifier, got {value!r}"
        )
    return value


def header_map(values: tuple[Any, ...]) -> dict[str, int]:
    headers = {str(value).strip(): index for index, value in enumerate(values) if value is not None}
    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        missing_text = ", ".join(missing)
        raise CatalogError(f"missing required workbook columns: {missing_text}")
    return headers


def find_header_row(worksheet: Any) -> tuple[int, dict[str, int]]:
    for row_number, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if not values or not any(value is not None for value in values):
            continue
        if all(header in values for header in REQUIRED_HEADERS):
            return row_number, header_map(tuple(values))
    raise CatalogError(
        f"sheet {worksheet.title!r}: no header row with the required approved-model columns"
    )


def profile_from_row(
    values: tuple[Any, ...],
    columns: dict[str, int],
    row_number: int,
) -> ModelProfile | None:
    if not any(value is not None for value in values):
        return None

    def value(header: str) -> Any:
        index = columns[header]
        return values[index] if index < len(values) else None

    provider = validate_identifier(
        clean_text(value("Provider"), field="Provider", row=row_number).casefold(),
        field="Provider",
        row=row_number,
    )
    model = validate_identifier(
        clean_text(value("Model"), field="Model", row=row_number).casefold(),
        field="Model",
        row=row_number,
    )
    thinking = clean_text(value("Thinking"), field="Thinking", row=row_number).casefold()
    if thinking not in THINKING_LEVELS:
        allowed = ", ".join(sorted(THINKING_LEVELS))
        raise CatalogError(f"row {row_number}: Thinking must be one of {allowed}")

    return ModelProfile(
        provider=provider,
        model=model,
        thinking=thinking,
        cost_per_task=positive_number(
            value("Cost/task (USD)"), field="Cost/task (USD)", row=row_number
        ),
        time_per_task=positive_number(
            value("Time/task (min)"), field="Time/task (min)", row=row_number
        ),
        intelligence_index=positive_number(
            value("Intelligence index"), field="Intelligence index", row=row_number
        ),
    )


def load_profiles(workbook_path: Path) -> tuple[ModelProfile, ...]:
    """Load and validate approved profiles from the first matching worksheet."""

    if not workbook_path.is_file():
        raise CatalogError(f"approved model workbook does not exist: {workbook_path}")

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        raise CatalogError(f"cannot read workbook {workbook_path}: {exc}") from exc

    try:
        for worksheet in workbook.worksheets:
            try:
                header_row, columns = find_header_row(worksheet)
            except CatalogError:
                continue

            profiles: list[ModelProfile] = []
            for row_number, values in enumerate(
                worksheet.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                profile = profile_from_row(tuple(values), columns, row_number)
                if profile is not None:
                    profiles.append(profile)

            if not profiles:
                raise CatalogError(f"sheet {worksheet.title!r}: no approved model rows found")

            keys = [(profile.provider, profile.model, profile.thinking) for profile in profiles]
            duplicates = sorted({key for key in keys if keys.count(key) > 1})
            if duplicates:
                raise CatalogError(f"duplicate approved model profiles: {duplicates}")
            return tuple(profiles)
    finally:
        workbook.close()

    raise CatalogError("no worksheet contains the required approved-model columns")


def display_number(value: float) -> str:
    return f"{value:.6f}".rstrip("0").rstrip(".")


def catalog_text(
    workbook_path: Path,
    profiles: tuple[ModelProfile, ...],
    project_root: Path,
) -> str:
    """Render the deterministic catalog consumed by the model-chooser skill."""

    try:
        display_workbook = workbook_path.resolve().relative_to(project_root.resolve()).as_posix()
    except ValueError:
        display_workbook = workbook_path.as_posix()

    lines = [
        "<!-- Generated by .agents/tools/model-chooser/catalog.py; do not edit manually. -->",
        "# Approved delegation model catalog",
        "",
        "Presence in this catalog means the profile is approved for delegation.",
        "Metrics are approximate comparative measures from Artificial Analysis:",
        f"{ARTIFICIAL_ANALYSIS_URL}",
        "The workbook is the reviewable source; this file is a generated runtime snapshot.",
        "",
        f"Source workbook: `{display_workbook}`",
        "",
        "| Provider | Model | Thinking | Cost/task (USD) | Time/task (min) | "
        "Intelligence index |",
        "| --- | --- | --- | ---: | ---: | ---: |",
    ]
    for profile in profiles:
        lines.append(
            "| "
            + " | ".join(
                (
                    profile.provider,
                    profile.model,
                    profile.thinking,
                    display_number(profile.cost_per_task),
                    display_number(profile.time_per_task),
                    display_number(profile.intelligence_index),
                )
            )
            + " |"
        )

    lines.extend(
        [
            "",
            "Selection uses the task's quality, correctness, speed, cost, complexity, and",
            "reversibility needs. Lower Time/task is faster; higher Intelligence index is",
            "stronger. Preserve the exact Provider, Model, and Thinking values.",
            "",
        ]
    )
    return "\n".join(lines)


def project_path(project_root: Path, value: Path | None, default_name: str) -> Path:
    path = value if value is not None else DEFAULT_DATA_DIR / default_name
    return path if path.is_absolute() else project_root / path


def generate(project_root: Path, workbook: Path | None, output: Path | None) -> int:
    workbook_path = project_path(project_root, workbook, DEFAULT_WORKBOOK)
    output_path = project_path(project_root, output, DEFAULT_CATALOG)
    profiles = load_profiles(workbook_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        catalog_text(workbook_path, profiles, project_root), encoding="utf-8"
    )
    print(f"Generated {output_path} from {workbook_path} ({len(profiles)} approved profiles)")
    return 0


def check(project_root: Path, workbook: Path | None, output: Path | None) -> int:
    workbook_path = project_path(project_root, workbook, DEFAULT_WORKBOOK)
    output_path = project_path(project_root, output, DEFAULT_CATALOG)
    expected = catalog_text(
        workbook_path, load_profiles(workbook_path), project_root
    )
    if not output_path.is_file():
        print(f"Catalog is missing: {output_path}", file=sys.stderr)
        return 1
    if output_path.read_text(encoding="utf-8") != expected:
        print(f"Catalog is out of date: {output_path}", file=sys.stderr)
        return 1
    print(f"Catalog is valid and up to date: {output_path}")
    return 0


app = App(help="Validate approved Excel model profiles and generate the runtime catalog.")


@app.command(name="generate")
def generate_command(
    *,
    project_root: Annotated[
        Path, Parameter(help="Repository root containing .agents/model-choosing.")
    ] = DEFAULT_PROJECT_ROOT,
    workbook: Annotated[
        Path | None,
        Parameter(
            help="Approved workbook; defaults to .agents/model-choosing/model-profiles.xlsx."
        ),
    ] = None,
    output: Annotated[
        Path | None,
        Parameter(
            help="Catalog output; defaults to .agents/model-choosing/model-catalog.md."
        ),
    ] = None,
) -> None:
    """Validate the approved workbook and generate model-catalog.md."""

    try:
        generate(project_root, workbook, output)
    except CatalogError as exc:
        raise SystemExit(str(exc)) from exc


@app.command(name="check")
def check_command(
    *,
    project_root: Annotated[
        Path, Parameter(help="Repository root containing .agents/model-choosing.")
    ] = DEFAULT_PROJECT_ROOT,
    workbook: Annotated[
        Path | None,
        Parameter(
            help="Approved workbook; defaults to .agents/model-choosing/model-profiles.xlsx."
        ),
    ] = None,
    output: Annotated[
        Path | None,
        Parameter(
            help="Catalog output; defaults to .agents/model-choosing/model-catalog.md."
        ),
    ] = None,
) -> None:
    """Validate the workbook and verify that the generated catalog is current."""

    try:
        raise SystemExit(check(project_root, workbook, output))
    except CatalogError as exc:
        raise SystemExit(str(exc)) from exc


if __name__ == "__main__":
    app()
